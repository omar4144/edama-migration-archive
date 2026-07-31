"""
Build a single Excel workbook for human review of the audit deliverables.
READ-ONLY: only reads CSV/MD reports under /app/memory/ and writes the .xlsx.

Sheets (Arabic RTL headers) — updated in Iteration 13.1:
  1. الملخص
  2. السجل الموحد ١١٨
  3. المشاركات الموحدة ١١٨
  4. مرشحو عبر الدفعات
  5. جودة Lovable 57
  6. سجل تدقيق مقترح
  7. صفوف المصادر ١٧٥
  8. مجموعات التطابق
  9. سجل ١٧٥ الخام
 10. الجمعيات متعددة الدفعات
 11. قرارات بشرية مطلوبة
"""
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MEMORY = Path("/app/memory")
OUT = MEMORY / "PARTICIPATING_ORGANIZATIONS_REVIEW.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0F3D3E")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Tahoma", size=11)
BODY_FONT = Font(name="Tahoma", size=10)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
HUMAN_FILL = PatternFill("solid", fgColor="FCE4D6")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
INFO_FILL = PatternFill("solid", fgColor="DDEBF7")


def _read_csv(path: Path):
    with path.open("r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames or []
    return headers, rows


def _write_sheet(ws, headers, rows, *, row_fill=None, freeze=True):
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            v = row.get(h, "") if isinstance(row, dict) else row[c_idx - 1]
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        if row_fill:
            fill = row_fill(row) if callable(row_fill) else row_fill
            if fill:
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill
    for c_idx, h in enumerate(headers, start=1):
        max_len = min(60, max(12, len(str(h)) + 2))
        for row in rows[:200]:
            v = row.get(h, "") if isinstance(row, dict) else row[c_idx - 1]
            L = len(str(v).splitlines()[0]) if v else 0
            if L > max_len:
                max_len = min(60, L + 2)
        ws.column_dimensions[get_column_letter(c_idx)].width = max_len
    ws.row_dimensions[1].height = 32
    if freeze:
        ws.freeze_panes = "A2"


def build_summary(wb, counts):
    ws = wb.create_sheet("الملخص", 0)
    ws.sheet_view.rightToLeft = True

    lines = [
        ("التدقيق الموحد — Iteration 13.1", True),
        ("قراءة فقط. لا تعديل ولا دمج ولا حذف في قاعدة البيانات.", False),
        ("", False),
        ("القرارات المُعتمَدة (بناءً على مراجعة المالك)", True),
        ("• 118 جمعية موحدة هي العدد الرسمي.", False),
        ("• الـ175 صف مصدر داخلي فقط، لا تُعرض في المنصة كعدد جمعيات أو مشاركات.", False),
        ("• دمج «صندوق الشهداء» ↔ «صندوق الشهداء والمصابين والأسرى والمفقودين» تمّت الموافقة عليه.", False),
        ("• «مقبول» في Lovable = LINK_EXISTS_CONTENT_NOT_VERIFIED (ليس دليل تخرّج).", False),
        ("", False),
        ("المؤشرات الرئيسية", True),
        (f"• صفوف مصادر الجهات: {counts['source_records']}", False),
        (f"• جمعيات موحدة: {counts['unified_orgs']}", False),
        (f"• مشاركات موحدة (org × cohort): {counts['unified_parts']}", False),
        (f"• UNIFIED_EXACT: {counts['n_exact']}", False),
        (f"• UNIFIED_PROBABLE_HUMAN_APPROVED: {counts['n_probable']}", False),
        (f"• LEGACY_ONLY: {counts['n_legacy_only']}", False),
        (f"• LOVABLE_ONLY: {counts['n_lovable_only']}", False),
        ("", False),
        ("توزيع الدفعات بعد التوحيد", True),
        (f"• الدفعة الأولى: {counts['coh1']}", False),
        (f"• الدفعة الثانية: {counts['coh2']}", False),
        (f"• الدفعة الثالثة: {counts['coh3']}", False),
        (f"• الدفعة الرابعة: {counts['coh4']}", False),
        (f"• الإجمالي: {counts['coh1']+counts['coh2']+counts['coh3']+counts['coh4']}", False),
        ("", False),
        ("مرشحو التطابق عبر الدفعات", True),
        (f"• عدد الأزواج فوق العتبة: {counts['xc']}", False),
        ("• لن تُدمج تلقائيًا — قرار بشري فقط.", False),
        ("• حالة تستحق تحققًا: «جمعية حوائج لحفظ النعمة» (دفعة 2) ↔ «جمعية حفظ النعمة» (دفعة 3).", False),
        ("", False),
        ("جودة الـ57 جهة Lovable", True),
        (f"• الجهات: {counts['lov_q']} — كلها LINK_EXISTS_CONTENT_NOT_VERIFIED.", False),
        ("• لم يُفحص محتوى أي ملف Google. لا يوجد ادعاء تخرّج.", False),
        ("", False),
        ("تأكيد سلامة البيانات", True),
        ("• لم تُعدَّل أي مجموعة في قاعدة البيانات.", False),
        ("• الـ175 صف الأصلية محفوظة في ORGANIZATION_PARTICIPATION_SOURCE_RECORDS.csv كطبقة أدلة.", False),
        ("• لا يبدأ Iteration 12 حتى تعتمد المخرجات.", False),
    ]

    for i, (txt, is_h) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        if is_h:
            c.font = Font(bold=True, name="Tahoma", size=13, color="0F3D3E")
            c.fill = PatternFill("solid", fgColor="E7F3F2")
        else:
            c.font = Font(name="Tahoma", size=11)
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 110


def build_multi_cohort(wb, part_rows):
    ws = wb.create_sheet("الجمعيات متعددة الدفعات")
    by_org = defaultdict(list)
    for r in part_rows:
        by_org[r.get("proposed_canonical_org_id", "")].append(r)

    headers = [
        "proposed_canonical_org_id", "canonical_org_name", "cohorts_present",
        "cohort_count", "participation_ids", "sources", "requires_review",
    ]
    rows = []
    for oid, group in by_org.items():
        cohorts = sorted({str(g.get("cohort", "")).strip() for g in group if g.get("cohort")})
        if len(cohorts) < 2:
            continue
        rows.append({
            "proposed_canonical_org_id": oid,
            "canonical_org_name": group[0].get("canonical_org_name", ""),
            "cohorts_present": " | ".join(cohorts),
            "cohort_count": len(cohorts),
            "participation_ids": " | ".join(g.get("participation_id", "") for g in group),
            "sources": " | ".join(sorted({g.get("source", "") for g in group})),
            "requires_review": " | ".join(g.get("requires_review", "") for g in group),
        })
    if not rows:
        _write_sheet(ws, headers, [{
            "proposed_canonical_org_id": "—",
            "canonical_org_name": "لا توجد جمعيات موحدة ظهرت في أكثر من دفعة.",
            "cohorts_present": "",
            "cohort_count": 0,
            "participation_ids": "",
            "sources": "",
            "requires_review": "",
        }], row_fill=GOOD_FILL)
    else:
        _write_sheet(ws, headers, rows, row_fill=WARN_FILL)


def build_human_decisions(wb, rows175, groups_rows, xc_rows, audit_rows):
    ws = wb.create_sheet("قرارات بشرية مطلوبة")
    headers = [
        "case_type", "candidate_or_group_id", "display_name",
        "match_status", "match_reason", "confidence",
        "why_needs_human", "recommended_action", "linked_records",
    ]
    rows = []

    # Proposed audit entries (approved but not applied)
    for a in audit_rows:
        rows.append({
            "case_type": "قرار موحد مقترح للتطبيق",
            "candidate_or_group_id": a.get("proposed_entry_id", ""),
            "display_name": a.get("canonical_org_name", ""),
            "match_status": a.get("action_type", ""),
            "match_reason": a.get("reason_code", ""),
            "confidence": a.get("similarity_score", ""),
            "why_needs_human": "تمّت الموافقة بشريًا — بانتظار تنفيذ سجل التدقيق.",
            "recommended_action": "تطبيق الدمج في participating_orgs + crosswalk_organizations",
            "linked_records": a.get("member_ids", ""),
        })

    # Top cross-cohort candidates
    for x in xc_rows[:50]:
        rows.append({
            "case_type": "مرشح تطابق عبر الدفعات",
            "candidate_or_group_id": x.get("candidate_pair_id", ""),
            "display_name": f"{x.get('org_a_name', '')} ↔ {x.get('org_b_name', '')}",
            "match_status": f"cohort {x.get('org_a_cohorts', '')} vs {x.get('org_b_cohorts', '')}",
            "match_reason": x.get("similarity_reason", ""),
            "confidence": f"ratio={x.get('similarity_ratio','')} · jaccard={x.get('jaccard','')}",
            "why_needs_human": "أسماء متشابهة في دفعات مختلفة — يحتاج قرار بشري (لا دمج تلقائي).",
            "recommended_action": "تحقق يدوي: هل هي نفس الجهة أم جهتان مختلفتان؟",
            "linked_records": f"{x.get('org_a_canonical_id','')} ↔ {x.get('org_b_canonical_id','')}",
        })

    # LEGACY_ONLY without Lovable
    for r in rows175:
        if r.get("organization_crosswalk_status") in ("LEGACY_ONLY", "NO_MATCH_LEGACY_ONLY"):
            rows.append({
                "case_type": "Legacy بدون مطابق Lovable",
                "candidate_or_group_id": r.get("registry_candidate_id", ""),
                "display_name": r.get("display_name", ""),
                "match_status": r.get("organization_crosswalk_status", ""),
                "match_reason": "لا يوجد صف Lovable مطابق",
                "confidence": r.get("match_confidence", ""),
                "why_needs_human": "تأكيد ما إذا كانت الجمعية لم تعد نشطة أم تحتاج إنشاء ORG في Lovable",
                "recommended_action": "إبقاء كـ Legacy فقط، أو ربطها يدويًا بـ ORG جديد",
                "linked_records": r.get("legacy_org_id", ""),
            })

    _write_sheet(ws, headers, rows, row_fill=HUMAN_FILL)


def main():
    hdr_src, src_rows = _read_csv(MEMORY / "ORGANIZATION_PARTICIPATION_SOURCE_RECORDS.csv")
    hdr_part_u, part_u_rows = _read_csv(MEMORY / "ORGANIZATION_COHORT_PARTICIPATIONS_UNIFIED.csv")
    hdr_reg, reg_rows = _read_csv(MEMORY / "ORGANIZATION_UNIFIED_REGISTRY.csv")
    hdr_xc, xc_rows = _read_csv(MEMORY / "CROSS_COHORT_CANDIDATES.csv")
    hdr_lovq, lovq_rows = _read_csv(MEMORY / "LOVABLE_57_ORG_QUALITY.csv")
    hdr_audit, audit_rows = _read_csv(MEMORY / "PROPOSED_AUDIT_LOG_ENTRIES.csv")

    hdr_175, rows_175 = _read_csv(MEMORY / "PARTICIPATING_ORGANIZATIONS_175_AUDIT.csv")
    hdr_grp, grp_rows = _read_csv(MEMORY / "ORGANIZATION_MATCH_GROUPS.csv")
    hdr_part, part_rows = _read_csv(MEMORY / "ORGANIZATION_COHORT_PARTICIPATIONS.csv")

    counts = {
        "source_records": len(src_rows),
        "unified_orgs": len(reg_rows),
        "unified_parts": len(part_u_rows),
        "n_exact": sum(1 for r in reg_rows if r["unification_status"] == "UNIFIED_EXACT"),
        "n_probable": sum(1 for r in reg_rows if r["unification_status"] == "UNIFIED_PROBABLE_HUMAN_APPROVED"),
        "n_legacy_only": sum(1 for r in reg_rows if r["unification_status"] == "LEGACY_ONLY"),
        "n_lovable_only": sum(1 for r in reg_rows if r["unification_status"] == "LOVABLE_ONLY"),
        "coh1": sum(1 for r in part_u_rows if str(r["cohort"]) == "1"),
        "coh2": sum(1 for r in part_u_rows if str(r["cohort"]) == "2"),
        "coh3": sum(1 for r in part_u_rows if str(r["cohort"]) == "3"),
        "coh4": sum(1 for r in part_u_rows if str(r["cohort"]) == "4"),
        "xc": len([r for r in xc_rows if r.get("candidate_pair_id") not in ("", "-")]),
        "lov_q": len(lovq_rows),
    }

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Summary
    build_summary(wb, counts)

    # 2. Unified registry (118)
    ws2 = wb.create_sheet("السجل الموحد ١١٨")
    def _color_reg(row):
        st = row.get("unification_status", "")
        if st == "UNIFIED_EXACT":
            return GOOD_FILL
        if st == "UNIFIED_PROBABLE_HUMAN_APPROVED":
            return WARN_FILL
        if st == "LEGACY_ONLY":
            return HUMAN_FILL
        return INFO_FILL
    _write_sheet(ws2, hdr_reg, reg_rows, row_fill=_color_reg)

    # 3. Unified participations (118)
    ws3 = wb.create_sheet("المشاركات الموحدة ١١٨")
    def _color_partu(row):
        if str(row.get("requires_review", "")).lower() == "true":
            return HUMAN_FILL
        conf = (row.get("cohort_confidence") or "")
        if conf.startswith("HIGH"):
            return GOOD_FILL
        if conf.startswith("MEDIUM"):
            return INFO_FILL
        if conf.startswith("UNKNOWN"):
            return WARN_FILL
        return None
    _write_sheet(ws3, hdr_part_u, part_u_rows, row_fill=_color_partu)

    # 4. Cross-cohort candidates
    ws4 = wb.create_sheet("مرشحو عبر الدفعات")
    def _color_xc(row):
        try:
            r = float(row.get("similarity_ratio", 0))
        except (ValueError, TypeError):
            r = 0
        if r >= 0.9:
            return HUMAN_FILL
        if r >= 0.8:
            return WARN_FILL
        return INFO_FILL
    _write_sheet(ws4, hdr_xc, xc_rows, row_fill=_color_xc)

    # 5. Lovable quality
    ws5 = wb.create_sheet("جودة Lovable 57")
    def _color_lovq(row):
        if row.get("rows_match_expected") == "true":
            return WARN_FILL  # still not verified content — kept as warning
        return HUMAN_FILL
    _write_sheet(ws5, hdr_lovq, lovq_rows, row_fill=_color_lovq)

    # 6. Proposed audit log
    ws6 = wb.create_sheet("سجل تدقيق مقترح")
    _write_sheet(ws6, hdr_audit, audit_rows, row_fill=WARN_FILL)

    # 7. Source records (175 evidence)
    ws7 = wb.create_sheet("صفوف المصادر ١٧٥")
    def _color_src(row):
        return GOOD_FILL if row.get("source_side") == "lovable" else INFO_FILL
    _write_sheet(ws7, hdr_src, src_rows, row_fill=_color_src)

    # 8. Match groups (legacy from Iteration 13.0)
    ws8 = wb.create_sheet("مجموعات التطابق")
    def _color_grp(row):
        st = (row.get("match_status") or "").upper()
        if st == "EXACT_SAME_ORGANIZATION":
            return GOOD_FILL
        if st == "PROBABLE_NAME_VARIANT":
            return WARN_FILL
        return None
    _write_sheet(ws8, hdr_grp, grp_rows, row_fill=_color_grp)

    # 9. 175 raw candidates registry (legacy)
    ws9 = wb.create_sheet("سجل ١٧٥ الخام")
    def _color_175(row):
        st = (row.get("organization_crosswalk_status") or "").upper()
        if st == "EXACT_NORMALIZED":
            return GOOD_FILL
        if st == "PROBABLE_NAME_VARIANT":
            return WARN_FILL
        if st in ("LEGACY_ONLY", "NO_MATCH_LEGACY_ONLY"):
            return HUMAN_FILL
        return None
    _write_sheet(ws9, hdr_175, rows_175, row_fill=_color_175)

    # 10. Multi-cohort orgs (should be empty confirming no false claim)
    build_multi_cohort(wb, part_rows)

    # 11. Human decisions
    build_human_decisions(wb, rows_175, grp_rows, xc_rows, audit_rows)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  - {name}: {ws.max_row-1} صف بيانات × {ws.max_column} عمود")


if __name__ == "__main__":
    main()
